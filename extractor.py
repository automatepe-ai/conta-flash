"""
Motor de extracción SUNAT - Formulario 621 (IGV Renta Mensual)
Lógica pura sin dependencias de UI.
"""

import pdfplumber
import pandas as pd
import re
import zipfile
import tempfile
import shutil
from pathlib import Path
from io import BytesIO

# ============================================================
# DICCIONARIO DE CASILLAS SUNAT - FORMULARIO 621
# ============================================================
CASILLAS = {
    # --- IGV VENTAS ---
    "100": "Ventas Netas (Base)",
    "101": "IGV Ventas (Tributo)",
    "154": "Ventas Netas Gravadas Ley 31556 10% (Base)",
    "155": "IGV 10% Ley 31556 (Tributo)",
    "160": "Ventas bienes Ley 27037 (Base)",
    "161": "Ventas bienes Ley 27037 (Tributo)",
    "162": "Descuentos y devoluciones Ley 27037 (Base)",
    "163": "Descuentos y devoluciones Ley 27037 (Tributo)",
    "106": "Exportaciones Facturadas en el período",
    "127": "Exportaciones Embarcadas en el período",
    "105": "Ventas no Gravadas",
    "109": "Ventas no Gravadas (Sin efecto en ratio)",
    "112": "Otras ventas",
    "131": "Total Tributo Bruto",
    # --- IGV COMPRAS ---
    "107": "Compras Netas Nacionales (Base)",
    "108": "IGV Compras Nacionales (Tributo)",
    "110": "Compras Nac. gravadas y no gravadas (Base)",
    "111": "Compras Nac. gravadas y no gravadas (Tributo)",
    "113": "Compras Nac. destinadas a no gravadas",
    "102": "Descuentos Concedidos (Base)",
    "103": "Descuentos Concedidos (Tributo)",
    "156": "Compras Netas tasa 10% Ley 31556 (Base)",
    "157": "Compras Netas tasa 10% Ley 31556 (Tributo)",
    "114": "Compras Importadas gravadas (Base)",
    "115": "Compras Importadas gravadas (Tributo)",
    "116": "Compras Importadas gravadas y no gravadas (Base)",
    "117": "Compras Importadas gravadas y no gravadas (Tributo)",
    "119": "Compras Importadas no gravadas exclusivamente",
    "120": "Compras internas no gravadas",
    "122": "Compras importadas no gravadas",
    "178": "Total Crédito Fiscal",
    "172": "Crédito Fiscal Especial",
    "173": "Coeficiente Crédito Fiscal",
    # --- IVAP ---
    "340": "Ventas Gravadas IVAP (Base)",
    "341": "Ventas Gravadas IVAP (Tributo)",
    "182": "Otros Créditos IVAP",
    # --- RENTA ---
    "301": "Ingresos Netos (Base)",
    "312": "Impuesto Renta (Tributo)",
    "380": "Coeficiente Renta",
    "315": "Porcentaje Renta",
    "336": "Pagos a Cuenta en Exceso",
    # --- DETERMINACIÓN DE LA DEUDA ---
    "140": "Impuesto Resultante IGV",
    "353": "Impuesto Resultante IVAP",
    "302": "Impuesto Resultante Renta",
    "145": "Saldo a Favor Período Anterior IGV",
    "351": "Saldo a Favor Período Anterior IVAP",
    "303": "Saldo a Favor Período Anterior Renta",
    "184": "Tributo a Pagar o Saldo a Favor IGV",
    "352": "Tributo a Pagar o Saldo a Favor IVAP",
    "304": "Tributo a Pagar o Saldo a Favor Renta",
    "171": "Percepciones declaradas en el período",
    "168": "Saldo de percepciones anteriores",
    "164": "Saldo de percepciones no aplicadas",
    "179": "Retenciones declaradas en el período",
    "176": "Saldo de retenciones anteriores",
    "165": "Saldo de Retenciones no aplicadas",
    "326": "Retenciones de tercera declaradas en período",
    "327": "Retenciones de tercera períodos anteriores",
    "347": "Compensación Saldo Exportador IGV",
    "305": "Compensación Saldo Exportador Renta",
    "328": "Impuesto Temporal Activos Netos (ITAN)",
    "681": "Sub Total IGV",
    "683": "Sub Total IVAP",
    "682": "Sub Total Renta",
    "185": "Pagos previos IGV",
    "342": "Pagos previos IVAP",
    "317": "Pagos previos Renta",
    "187": "Interés moratorio IGV",
    "343": "Interés moratorio IVAP",
    "319": "Interés moratorio Renta",
    "188": "Total deuda tributaria IGV",
    "344": "Total deuda tributaria IVAP",
    "324": "Total deuda tributaria Renta",
}


# ============================================================
# FUNCIONES DE EXTRACCIÓN
# ============================================================
def extract_header(full_text: str) -> dict:
    """Extrae campos del encabezado del formulario 621."""
    header = {}

    ruc_match = re.search(r'RUC\s+(\d{11})', full_text)
    header['RUC'] = ruc_match.group(1) if ruc_match else ''

    razon_match = re.search(r'Raz[oó]n\s+Social\s+(.+?)\s+Per[ií]odo', full_text)
    header['Razon_Social'] = razon_match.group(1).strip() if razon_match else ''

    periodo_match = re.search(r'Per[ií]odo\s+(\d{6})', full_text)
    header['Periodo'] = periodo_match.group(1) if periodo_match else ''

    fecha_match = re.search(r'Fecha\s+de\s+Presentaci[oó]n\s+(\d{2}/\d{2}/\d{4})', full_text)
    header['Fecha_Presentacion'] = fecha_match.group(1) if fecha_match else ''

    tipo_decl_match = re.search(r'Tipo\s+de\s+Declaraci[oó]n\s+(\w+)', full_text)
    header['Tipo_Declaracion'] = tipo_decl_match.group(1) if tipo_decl_match else ''

    moneda_match = re.search(r'Tipo\s+de\s+Moneda\s+(\w+)', full_text)
    header['Tipo_Moneda'] = moneda_match.group(1) if moneda_match else ''

    orden_match = re.search(r'N[uú]mero\s+de\s+Orden\s+(\d+)', full_text)
    header['Numero_Orden'] = orden_match.group(1) if orden_match else ''

    return header


def extract_casillas(full_text: str) -> dict:
    """Extrae valores de TODAS las casillas del formulario."""
    casilla_values = {}
    for code in CASILLAS:
        pattern = rf'\b{code}\b\s+(-?[\d,]+\.\d{{1,2}})'
        match = re.search(pattern, full_text)
        if match:
            value_str = match.group(1).replace(',', '')
            casilla_values[code] = float(value_str)
        else:
            casilla_values[code] = None
    return casilla_values


def extract_pdf_bytes(pdf_bytes: bytes, filename: str) -> dict:
    """Extrae datos de un PDF en memoria (bytes)."""
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        full_text = '\n'.join(
            page.extract_text() or '' for page in pdf.pages
        )
    header = extract_header(full_text)
    casillas = extract_casillas(full_text)
    record = {**header, **{f'C{k}': v for k, v in casillas.items()}}
    record['Archivo_Origen'] = filename
    return record


def _build_description_row() -> dict:
    """Fila de descripciones para las casillas."""
    desc_row = {}
    desc_row['Archivo_Origen'] = 'DESCRIPCIÓN'
    for code, desc in CASILLAS.items():
        desc_row[f'C{code}'] = desc
    return desc_row


def _order_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ordena columnas: header + casillas numéricas."""
    header_cols = ['Archivo_Origen', 'RUC', 'Razon_Social', 'Periodo',
                   'Fecha_Presentacion', 'Tipo_Declaracion', 'Tipo_Moneda', 'Numero_Orden']
    casilla_cols = sorted(
        [c for c in df.columns if c.startswith('C') and c[1:].isdigit()],
        key=lambda x: int(x[1:])
    )
    for col in header_cols:
        if col not in df.columns:
            df[col] = ''
    return df[header_cols + casilla_cols].sort_values('Periodo').reset_index(drop=True)


def _apply_excel_styles(ws, df):
    """Aplica estilos profesionales al worksheet de openpyxl."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    if ws.max_row < 1 or ws.max_column < 1:
        return

    HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    DESC_FILL = PatternFill(start_color="F5F0E8", end_color="F5F0E8", fill_type="solid")
    DESC_FONT = Font(italic=True, color="6B6560", size=9)
    BORDER = Border(
        left=Side(style="thin", color="E8E0D6"),
        right=Side(style="thin", color="E8E0D6"),
        top=Side(style="thin", color="E8E0D6"),
        bottom=Side(style="thin", color="E8E0D6"),
    )
    ALT_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

    # Headers (fila 1)
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Descripción (fila 2)
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = DESC_FONT
        cell.fill = DESC_FILL
        cell.border = BORDER

    # Datos (filas 3+)
    for row_idx in range(3, len(df) + 3):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
            col_name = df.columns[col_idx - 1]
            if col_name.startswith('C') and col_name[1:].isdigit():
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Congelar paneles
    if len(df.columns) >= 2:
        ws.freeze_panes = "C3"


def _write_excel_bytes(df: pd.DataFrame, desc_row: dict, include_dict: bool = True,
                       is_pro: bool = True, free_period_limit: int = 3) -> bytes:
    """Genera un Excel en memoria y retorna los bytes."""
    from openpyxl.styles import Font

    # FILTRO FREE: Solo 3 períodos más recientes
    total_periodos = 0
    if not is_pro and 'Periodo' in df.columns:
        periodos_unicos = sorted(df['Periodo'].dropna().unique())
        total_periodos = len(periodos_unicos)
        if total_periodos > free_period_limit:
            periodos_permitidos = periodos_unicos[-free_period_limit:]
            df = df[df['Periodo'].isin(periodos_permitidos)]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Consolidado', index=False, startrow=1)
        ws = writer.sheets['Consolidado']
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            if col_name in desc_row:
                ws.cell(row=2, column=col_idx, value=desc_row[col_name])
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), 12)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        _apply_excel_styles(ws, df)

        # AVISO FREE: Agregar fila informativa al final
        if not is_pro and total_periodos > free_period_limit:
            last_row = len(df) + 3
            aviso_text = (f"⚠️ MODO GRATUITO: Solo se muestran {free_period_limit} de {total_periodos} períodos. "
                         f"Para ver todos los períodos, activa Plan Pro.")
            ws.cell(row=last_row, column=1, value=aviso_text)
            ws.cell(row=last_row, column=1).font = Font(bold=True, color="FF6600", size=10)
            ws.merge_cells(start_row=last_row, start_column=1,
                          end_row=last_row, end_column=min(5, len(df.columns)))

        if include_dict:
            dict_df = pd.DataFrame([
                {'Casilla': code, 'Descripcion': desc}
                for code, desc in sorted(CASILLAS.items(), key=lambda x: int(x[0]))
            ])
            dict_df.to_excel(writer, sheet_name='Diccionario_Casillas', index=False)

        # Sheet de aviso para FREE
        if not is_pro:
            aviso_df = pd.DataFrame([{
                'Aviso': '⚠️ PLAN GRATUITO - ContaFlash',
                'Detalle': (f'Esta exportación contiene solo {free_period_limit} de {total_periodos} períodos. '
                           f'Para acceso completo, contacta: https://wa.me/51962927872')
            }])
            aviso_df.to_excel(writer, sheet_name='Aviso_Plan_Gratuito', index=False)

    return output.getvalue()


# ============================================================
# VALIDACIONES AUTOMÁTICAS
# ============================================================
def validate_dataframe(df: pd.DataFrame) -> list:
    """
    Valida el DataFrame procesado y retorna lista de advertencias.
    Cada advertencia es un dict: {'tipo': str, 'mensaje': str, 'severidad': str}
    """
    warnings = []

    if df is None or len(df) == 0:
        return warnings

    # 1. Declaraciones duplicadas (mismo RUC + mismo Período)
    if 'RUC' in df.columns and 'Periodo' in df.columns:
        dupes = df.groupby(['RUC', 'Periodo']).size().reset_index(name='count')
        dupes = dupes[dupes['count'] > 1]
        for _, row in dupes.iterrows():
            warnings.append({
                'tipo': 'duplicado',
                'mensaje': f"RUC {row['RUC']} tiene {row['count']} declaraciones para {row['Periodo']}",
                'severidad': 'warning',
            })

    # 2. RUCs duplicados en diferentes archivos
    if 'RUC' in df.columns and 'Archivo_Origen' in df.columns:
        ruc_files = df.groupby('RUC')['Archivo_Origen'].nunique().reset_index(name='n_archivos')
        ruc_files = ruc_files[ruc_files['n_archivos'] > 1]
        for _, row in ruc_files.iterrows():
            warnings.append({
                'tipo': 'ruc_duplicado',
                'mensaje': f"RUC {row['RUC']} aparece en {row['n_archivos']} archivos distintos",
                'severidad': 'info',
            })

    # 3. Información faltante (campos header vacíos)
    header_cols = ['RUC', 'Razon_Social', 'Periodo']
    for col in header_cols:
        if col in df.columns:
            missing = df[col].isna().sum() + (df[col] == '').sum()
            if missing > 0:
                warnings.append({
                    'tipo': 'faltante',
                    'mensaje': f"{missing} declaración(es) sin campo '{col}'",
                    'severidad': 'info',
                })

    # 4. Casillas con valores en 0 o vacíos (posibles errores de extracción)
    casilla_cols = [c for c in df.columns if c.startswith('C') and c[1:].isdigit()]
    key_casillas = ['C101', 'C108', 'C312']  # IGV Ventas, IGV Compras, Renta
    for casilla in key_casillas:
        if casilla in df.columns:
            zeros = (df[casilla] == 0).sum() + (df[casilla] == '').sum() + df[casilla].isna().sum()
            if zeros > 0 and zeros < len(df):
                desc = CASILLAS.get(casilla[1:], casilla)
                warnings.append({
                    'tipo': 'valor_cero',
                    'mensaje': f"{zeros} declaración(es) con {desc} en 0 o vacío",
                    'severidad': 'info',
                })

    return warnings


# ============================================================
# FUNCIÓN PRINCIPAL: Procesar PDFs desde archivos subidos
# ============================================================
def process_uploaded_pdfs(uploaded_files: list, is_pro: bool = True) -> tuple:
    """
    Procesa una lista de PDFs/ZIPs subidos.
    Retorna (excel_bytes, stats_dict, log_messages).
    """
    records = []
    errors = []
    logs = []

    for uploaded in uploaded_files:
        file_bytes = uploaded.read()
        filename = uploaded.name

        if filename.lower().endswith('.zip'):
            logs.append(f"📦 ZIP: {filename}")
            tmp_dir = tempfile.mkdtemp(prefix="sunat_web_")
            try:
                with zipfile.ZipFile(BytesIO(file_bytes), 'r') as zf:
                    zf.extractall(tmp_dir)
                for pdf_path in Path(tmp_dir).rglob('*.pdf'):
                    try:
                        pdf_bytes = pdf_path.read_bytes()
                        record = extract_pdf_bytes(pdf_bytes, pdf_path.name)
                        records.append(record)
                        logs.append(f"  ✅ {pdf_path.name} — Período {record.get('Periodo', '?')}")
                    except Exception as e:
                        errors.append({'archivo': pdf_path.name, 'error': str(e)})
                        logs.append(f"  ❌ {pdf_path.name} — {e}")
            except zipfile.BadZipFile:
                errors.append({'archivo': filename, 'error': 'ZIP corrupto'})
                logs.append(f"  ❌ ZIP corrupto")
            finally:
                shutil.rmtree(tmp_dir, ignore_errors=True)

        elif filename.lower().endswith('.pdf'):
            try:
                record = extract_pdf_bytes(file_bytes, filename)
                records.append(record)
                logs.append(f"✅ {filename} — Período {record.get('Periodo', '?')}")
            except Exception as e:
                errors.append({'archivo': filename, 'error': str(e)})
                logs.append(f"❌ {filename} — {e}")

    if not records:
        return None, {'total': 0, 'ok': 0, 'errors': len(errors)}, logs, None

    df = pd.DataFrame(records)
    df = _order_columns(df)
    desc_row = _build_description_row()
    excel_bytes = _write_excel_bytes(df, desc_row, is_pro=is_pro)

    # Calcular totales financieros
    def _safe_sum(col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    stats = {
        'total': len(records) + len(errors),
        'ok': len(records),
        'errors': len(errors),
        'periodos': sorted(df['Periodo'].dropna().unique().tolist()),
        'empresas': sorted(df['RUC'].dropna().unique().tolist()),
        'n_casillas': len([c for c in df.columns if c.startswith('C') and c[1:].isdigit()]),
        'warnings': validate_dataframe(df),
        'totales': {
            'igv_ventas': _safe_sum('C101'),
            'igv_compras': _safe_sum('C108'),
            'renta': _safe_sum('C312'),
        },
        'is_pro': is_pro,
    }
    return excel_bytes, stats, logs, df


# ============================================================
# FUNCIÓN PRINCIPAL: Consolidar Excels/CSVs subidos
# ============================================================
def _parse_uploaded_excel(file_bytes: bytes, filename: str, empresa: str) -> tuple:
    """Parsea un Excel/CSV del PDT 621 y retorna (record, error)."""
    try:
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(BytesIO(file_bytes), sep=';', dtype=str)
        else:
            df = pd.read_excel(BytesIO(file_bytes), dtype=str)

        df.columns = df.columns.str.strip()

        col_map = {}
        for col in df.columns:
            col_lower = col.lower().replace('_', ' ')
            if 'nro' in col_lower and 'casilla' in col_lower:
                col_map['casilla'] = col
            elif 'valor' in col_lower and 'casilla' in col_lower:
                col_map['valor'] = col
            elif col_lower == 'periodo':
                col_map['periodo'] = col
            elif 'ruc' in col_lower:
                col_map['ruc'] = col
            elif 'nro' in col_lower and 'orden' in col_lower:
                col_map['orden'] = col
            elif 'fecha' in col_lower and 'present' in col_lower:
                col_map['fecha'] = col
            elif 'rectif' in col_lower:
                col_map['rectif'] = col

        if 'casilla' not in col_map or 'valor' not in col_map:
            return None, None

        row0 = df.iloc[0] if len(df) > 0 else {}

        tipo_decl = 'Original'
        rectif_col = col_map.get('rectif', '')
        if rectif_col and str(row0.get(rectif_col, '0')).strip() not in ('0', '', 'nan'):
            tipo_decl = 'Rectificatoria'
        elif re.search(r'\.2\.\d{4}', filename) or filename.rsplit('.', 1)[0].endswith('.2'):
            tipo_decl = 'Rectificatoria'

        record = {
            'Empresa': empresa,
            'Archivo_Origen': filename,
            'RUC': str(row0.get(col_map.get('ruc', ''), '')),
            'Razon_Social': '',
            'Periodo': str(row0.get(col_map.get('periodo', ''), '')),
            'Fecha_Presentacion': str(row0.get(col_map.get('fecha', ''), '')),
            'Tipo_Declaracion': tipo_decl,
            'Tipo_Moneda': '',
            'Numero_Orden': str(row0.get(col_map.get('orden', ''), '')),
        }

        if not record['Periodo'] or record['Periodo'] == 'nan':
            match = re.search(r'(\d{6})', filename)
            record['Periodo'] = match.group(1) if match else ''

        for _, row in df.iterrows():
            casilla = str(row[col_map['casilla']]).strip()
            valor = row[col_map['valor']]
            try:
                record[f'C{casilla}'] = float(valor)
            except (ValueError, TypeError):
                record[f'C{casilla}'] = None

        return record, None
    except Exception as e:
        return None, {'empresa': empresa, 'archivo': filename, 'error': str(e)}


def consolidate_uploaded_excels(uploaded_files: list, empresa_name: str = "Mi Empresa",
                                is_pro: bool = True, free_period_limit: int = 3) -> tuple:
    """
    Consolida Excels/CSVs subidos del PDT 621.
    Retorna (excel_bytes, stats_dict, log_messages).
    """
    records = []
    errors = []
    logs = []

    for uploaded in uploaded_files:
        file_bytes = uploaded.read()
        filename = uploaded.name

        record, err = _parse_uploaded_excel(file_bytes, filename, empresa_name)
        if err:
            errors.append(err)
            logs.append(f"❌ {filename} — {err['error']}")
        elif record:
            tipo = record['Tipo_Declaracion']
            records.append(record)
            logs.append(f"✅ {filename} — {record['Periodo']} ({tipo})")
        else:
            logs.append(f"⏭️ {filename} — Sin columnas Casilla/Valor (omitido)")

    if not records:
        return None, {'total': 0, 'ok': 0, 'errors': len(errors)}, logs

    df = pd.DataFrame(records)

    # FILTRO FREE: Solo 3 períodos más recientes
    total_periodos = 0
    if not is_pro and 'Periodo' in df.columns:
        periodos_unicos = sorted(df['Periodo'].dropna().unique())
        total_periodos = len(periodos_unicos)
        if total_periodos > free_period_limit:
            periodos_permitidos = periodos_unicos[-free_period_limit:]
            df = df[df['Periodo'].isin(periodos_permitidos)]
            logs.append(f"ℹ️ Plan gratuito: Solo {free_period_limit} de {total_periodos} períodos incluidos")

    header_cols = ['Empresa', 'Archivo_Origen', 'RUC', 'Razon_Social', 'Periodo',
                   'Fecha_Presentacion', 'Tipo_Declaracion', 'Tipo_Moneda', 'Numero_Orden']
    casilla_cols = sorted(
        [c for c in df.columns if c.startswith('C') and c[1:].isdigit()],
        key=lambda x: int(x[1:])
    )
    for col in header_cols:
        if col not in df.columns:
            df[col] = ''

    df = df[header_cols + casilla_cols].sort_values(['Empresa', 'Periodo']).reset_index(drop=True)

    desc_row = {col: '' for col in df.columns}
    desc_row['Empresa'] = 'DESCRIPCIÓN'
    for code, desc in CASILLAS.items():
        col_name = f'C{code}'
        if col_name in desc_row:
            desc_row[col_name] = desc

    # Generar Excel con pestañas por empresa
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Resumen
        df.to_excel(writer, sheet_name='Resumen', index=False, startrow=1)
        ws = writer.sheets['Resumen']
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            if col_name in desc_row:
                ws.cell(row=2, column=col_idx, value=desc_row[col_name])
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), 12)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2
        _apply_excel_styles(ws, df)

        # Pestañas por empresa
        empresas = df['Empresa'].unique()
        for empresa in sorted(empresas):
            emp_df = df[df['Empresa'] == empresa].reset_index(drop=True)
            sheet_name = empresa[:31]
            emp_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(emp_df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
                if col_name in desc_row:
                    ws.cell(row=2, column=col_idx, value=desc_row[col_name])
            _apply_excel_styles(ws, emp_df)

        # Diccionario
        dict_df = pd.DataFrame([
            {'Casilla': code, 'Descripcion': desc}
            for code, desc in sorted(CASILLAS.items(), key=lambda x: int(x[0]))
        ])
        dict_df.to_excel(writer, sheet_name='Diccionario_Casillas', index=False)

        # Errores
        if errors:
            err_df = pd.DataFrame(errors)
            err_df.to_excel(writer, sheet_name='Errores', index=False)

        # AVISO FREE: Agregar sheet informativo
        if not is_pro and total_periodos > free_period_limit:
            aviso_df = pd.DataFrame([{
                'Aviso': '⚠️ PLAN GRATUITO - ContaFlash',
                'Detalle': (f'Esta consolidación contiene solo {free_period_limit} de {total_periodos} períodos. '
                           f'Para acceso completo, contacta: https://wa.me/51962927872')
            }])
            aviso_df.to_excel(writer, sheet_name='Aviso_Plan_Gratuito', index=False)

    def _safe_sum(col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    stats = {
        'total': len(records) + len(errors),
        'ok': len(records),
        'errors': len(errors),
        'empresas': len(empresas),
        'periodos': sorted(df['Periodo'].dropna().unique().tolist()),
        'periodos_total': total_periodos if not is_pro else len(df['Periodo'].dropna().unique()),
        'n_casillas': len(casilla_cols),
        'warnings': validate_dataframe(df),
        'totales': {
            'igv_ventas': _safe_sum('C101'),
            'igv_compras': _safe_sum('C108'),
            'renta': _safe_sum('C312'),
        },
        'is_pro': is_pro,
    }
    return output.getvalue(), stats, logs, df
